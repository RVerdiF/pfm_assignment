"""Tests for Step 1 - Excel -> DuckDB raw ingestion.

Behavioral focus:
- to_snake_case normalizes column names without breaking existing snake_case names.
- sheet_to_table maps sheet -> table name (tracknow_checkouts / posthog_sessions).
- load_excel_into_duckdb persists the raw schema + tables, without filtering/deduplicating
  real records (including fully null rows present in the extent).
- _read_sheet reads with Polars and retains fully null interior rows.
- validate performs reconciliation of rows/columns/names.
"""
import os

import duckdb
import openpyxl
import polars as pl
import pytest

from ingestion import load_excel as le

# ---------------------------------------------------------------------------
# Unit: to_snake_case
# ---------------------------------------------------------------------------

def test_snake_case_keeps_existing_snake_case():
    assert le.to_snake_case("click_id") == "click_id"
    assert le.to_snake_case("order_price_gbp") == "order_price_gbp"
    assert le.to_snake_case("utm_source") == "utm_source"
    assert le.to_snake_case("has_tracknow_conversion") == "has_tracknow_conversion"
    assert le.to_snake_case("affiliate_session_id") == "affiliate_session_id"


def test_snake_case_normalizes_spaces_and_punctuation():
    assert le.to_snake_case("Click ID") == "click_id"
    assert le.to_snake_case("Order Price (GBP)") == "order_price_gbp"
    assert le.to_snake_case("  Session Duration Seconds  ") == "session_duration_seconds"
    assert le.to_snake_case("utm content") == "utm_content"


def test_snake_case_splits_camel_boundaries():
    assert le.to_snake_case("hasCheckoutStarted") == "has_checkout_started"
    assert le.to_snake_case("tracknowUserId") == "tracknow_user_id"


# ---------------------------------------------------------------------------
# Unit: sheet_to_table
# ---------------------------------------------------------------------------

def test_sheet_to_table_maps_target_sheets():
    assert le.sheet_to_table("Sample TrackNow Checkouts") == "tracknow_checkouts"
    assert le.sheet_to_table("Sample PostHog Sessions") == "posthog_sessions"


# ---------------------------------------------------------------------------
# Integration: create small xlsx fixture and load into duckdb
# ---------------------------------------------------------------------------

def _write_fixture_xlsx(path):
    """Create sheets with real records and a fully null row IN THE MIDDLE of
    the extent (absolute row 4 with no cells) - represents a record present in
    the spreadsheet that ingestion must preserve, not a footer outside the data area."""
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Sample TrackNow Checkouts"
    _write_rows_at(ws1, {
        1: ["Click ID", "Order Price (GBP)"],   # header
        2: ["id-aaaa", 66.97],
        3: ["id-bbbb", None],
        # row 4: fully-null interior row (no cells written)
        5: ["id-cccc", 1.0],
    })

    ws2 = wb.create_sheet("Sample PostHog Sessions")
    _write_rows_at(ws2, {
        1: ["Session ID", "click_id_from_url", "utm_source"],  # header
        2: ["sess-1", "click-1", "google"],
        3: ["sess-2", None, "facebook"],
        # row 4: fully-null interior row (no cells written)
        5: ["sess-3", "click-3", None],
    })

    wb.save(path)


def _write_rows_at(ws, rows):
    """Write values at the given absolute rows (does not extend dimensions for
    rows without cells)."""
    for r_i, values in rows.items():
        for c_i, val in enumerate(values, start=1):
            ws.cell(row=r_i, column=c_i).value = val


@pytest.fixture()
def fixture_xlsx(tmp_path):
    p = tmp_path / "sample.xlsx"
    _write_fixture_xlsx(str(p))
    return str(p)


@pytest.fixture()
def duckdb_path(tmp_path):
    return str(tmp_path / "pfm.duckdb")


def test_load_creates_raw_tables(fixture_xlsx, duckdb_path):
    created = le.load_excel_into_duckdb(fixture_xlsx, duckdb_path)

    assert set(created) == {"tracknow_checkouts", "posthog_sessions"}

    con = duckdb.connect(duckdb_path)
    try:
        schemas = con.execute(
            "SELECT table_name FROM information_schema.tables WHERE table_schema='raw' ORDER BY table_name"
        ).fetchall()
        assert [r[0] for r in schemas] == ["posthog_sessions", "tracknow_checkouts"]
    finally:
        con.close()


def test_load_row_counts(fixture_xlsx, duckdb_path):
    """Counts match exactly the extent read from Excel (header + real records +
    interior fully null row): 4 rows per sheet."""
    le.load_excel_into_duckdb(fixture_xlsx, duckdb_path)
    con = duckdb.connect(duckdb_path)
    try:
        n_tracknow = con.execute(
            "SELECT count(*) FROM raw.tracknow_checkouts"
        ).fetchone()[0]
        n_posthog = con.execute(
            "SELECT count(*) FROM raw.posthog_sessions"
        ).fetchone()[0]
    finally:
        con.close()
    # 3 real records + 1 fully null row (interior of the extent)
    assert n_tracknow == 4
    assert n_posthog == 4


def test_load_applies_snake_case_to_columns(fixture_xlsx, duckdb_path):
    le.load_excel_into_duckdb(fixture_xlsx, duckdb_path)
    con = duckdb.connect(duckdb_path)
    try:
        cols1 = [r[0] for r in con.execute("DESCRIBE raw.tracknow_checkouts").fetchall()]
        cols2 = [r[0] for r in con.execute("DESCRIBE raw.posthog_sessions").fetchall()]
    finally:
        con.close()
    assert cols1 == ["click_id", "order_price_gbp"]
    assert cols2 == ["session_id", "click_id_from_url", "utm_source"]


def test_load_preserves_identifier_values(fixture_xlsx, duckdb_path):
    """Real identifiers are preserved; the fully null row is NOT interpreted
    as the end of the table - records following it exist."""
    le.load_excel_into_duckdb(fixture_xlsx, duckdb_path)
    con = duckdb.connect(duckdb_path)
    try:
        ids = con.execute(
            "SELECT click_id FROM raw.tracknow_checkouts WHERE click_id IS NOT NULL ORDER BY click_id"
        ).fetchall()
        sessions = con.execute(
            "SELECT session_id FROM raw.posthog_sessions WHERE session_id IS NOT NULL ORDER BY session_id"
        ).fetchall()
    finally:
        con.close()
    assert [r[0] for r in ids] == ["id-aaaa", "id-bbbb", "id-cccc"]
    assert [r[0] for r in sessions] == ["sess-1", "sess-2", "sess-3"]


def test_load_retains_fully_null_interior_rows(fixture_xlsx, duckdb_path):
    """A fully null row within the spreadsheet extent is a record and cannot
    be filtered out - no data filters are applied during ingestion."""
    le.load_excel_into_duckdb(fixture_xlsx, duckdb_path)
    con = duckdb.connect(duckdb_path)
    try:
        nulls = con.execute(
            "SELECT count(*) FROM raw.tracknow_checkouts "
            "WHERE click_id IS NULL AND order_price_gbp IS NULL"
        ).fetchone()[0]
        nulls_s = con.execute(
            "SELECT count(*) FROM raw.posthog_sessions "
            "WHERE session_id IS NULL AND click_id_from_url IS NULL AND utm_source IS NULL"
        ).fetchone()[0]
    finally:
        con.close()
    assert nulls == 1
    assert nulls_s == 1


def test_read_sheet_returns_polars_and_retains_null_rows(fixture_xlsx):
    """_read_sheet reads through Polars and keeps fully null interior rows
    (drop_empty_rows=False), so the DuckDB extent matches the source workbook."""
    df = le._read_sheet(fixture_xlsx, "Sample TrackNow Checkouts")
    assert isinstance(df, pl.DataFrame)
    assert df.height == 4
    assert df.columns == ["click_id", "order_price_gbp"]
    # The interior fully null row is present as a row of all nulls.
    assert df.filter(pl.col("click_id").is_null()).height == 1


def test_validate_reconciles_ok(fixture_xlsx, duckdb_path):
    le.load_excel_into_duckdb(fixture_xlsx, duckdb_path)
    report = le.validate(excel_path=fixture_xlsx, duckdb_path=duckdb_path)
    assert report["all_ok"] is True
    assert report["tables"]["tracknow_checkouts"]["rows"] == 4
    assert report["tables"]["tracknow_checkouts"]["cols"] == 2
    assert report["tables"]["posthog_sessions"]["rows"] == 4
    assert report["tables"]["posthog_sessions"]["cols"] == 3
    assert set(report["tables"]["tracknow_checkouts"]["columns"]) == {"click_id", "order_price_gbp"}
